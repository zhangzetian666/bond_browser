# -*- coding: utf-8 -*-
"""
债券数据浏览器 V3
支持债券类型筛选、央企独立板块、右键复制、分类明细展示。
"""
import json
import math
import os
import sys
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime

try:
    import openpyxl
except ImportError:
    messagebox.showerror("缺少依赖", "请先安装 openpyxl：pip install openpyxl")
    raise


def resource_path(relative_path):
    """兼容 PyInstaller 打包后的资源路径。"""
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(os.path.dirname(__file__) or ".")
    return os.path.join(base_path, relative_path)


# ---------------------------------------------------------------------------
# v5 数据列索引（模块级，便于 lambda 中引用）
# ---------------------------------------------------------------------------
COL_BOND_TYPE = 3          # 债券类型
COL_ISSUER = 5             # 发行人中文名称
COL_PROVINCE = 18          # 所属省
COL_CITY = 19              # 所属市
COL_DISTRICT = 20          # 所属区/县
COL_INDUSTRY = 21          # 申万一级行业
COL_ENTERPRISE_NATURE = 22 # 企业性质
COL_CENTRAL = 23           # 是否央企
COL_MUNICIPAL = 24         # 是否市级国企
COL_PROVINCIAL = 25        # 是否省级国企
COL_COUNTY = 26            # 是否县(区)级国企
COL_CONTROLLER = 27        # 实际控制人
COL_SHAREHOLDER = 28       # 股东
COL_LGFV_SUBSIDIARY = 30   # 是否城投子公司
COL_LGFV_LEVEL = 31        # 城投平台层级
COL_BOND_COUNT = 32        # 债券存量只数
COL_TOTAL_ASSETS = 33      # 资产总计
COL_TOTAL_EQUITY = 34      # 所有者权益合计
COL_TOTAL_REVENUE = 35     # 营业总收入
COL_NET_PROFIT = 36        # 净利润
COL_CASH_FLOW_2024 = 37    # 经营活动现金流 2024
COL_CASH_FLOW_2025 = 38    # 经营活动现金流 2025
COL_ROA = 39               # 总资产报酬率(%)
COL_REVENUE_RATIO = 42     # 营业收入占比(%)
COL_REVENUE = 43           # 营业收入
COL_PROJECT_NAME = 44      # 项目名称

COL_SEQ = 0                # 序号
COL_BOND_SHORT = 1         # 债券简称
COL_BOND_CODE = 2          # 债券代码
COL_BOND_FULL = 4          # 债券全称
COL_ISSUE_AMOUNT = 6       # 发行总额
COL_BOND_TERM = 7          # 债券期限
COL_START_DATE = 8         # 起息日期
COL_COUPON_RATE = 9        # 发行票面利率
COL_USE_OF_PROCEEDS = 13   # 募集资金用途
COL_BOND_RATING = 14       # 发行时债券评级
COL_ISSUER_RATING = 15     # 发行时主体评级
COL_GUARANTOR = 16         # 担保人名称
COL_GUARANTOR_RATING = 17  # 最新担保人评级
COL_LISTING_EXCHANGE = 40  # 上市交易所


class ScrollableButtonGrid(ttk.Frame):
    """可垂直滚动的按钮网格。"""

    def __init__(self, parent, cols=5, *args, **kwargs):
        super().__init__(parent, *args, **kwargs)
        self.cols = cols

        self.rowconfigure(0, weight=1)
        self.columnconfigure(0, weight=1)

        self.canvas = tk.Canvas(self, highlightthickness=0)
        self.canvas.grid(row=0, column=0, sticky="nsew")

        self.vsb = ttk.Scrollbar(self, orient=tk.VERTICAL, command=self.canvas.yview)
        self.vsb.grid(row=0, column=1, sticky="ns")
        self.canvas.configure(yscrollcommand=self.vsb.set)

        self.inner = ttk.Frame(self.canvas)
        self.canvas_window = self.canvas.create_window((0, 0), window=self.inner, anchor="nw")

        self.inner.bind("<Configure>", self._on_inner_configure)
        self.canvas.bind("<Configure>", self._on_canvas_configure)
        self.canvas.bind("<MouseWheel>", self._on_mousewheel)

    def _on_inner_configure(self, event=None):
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))

    def _on_canvas_configure(self, event=None):
        width = self.canvas.winfo_width()
        self.canvas.itemconfig(self.canvas_window, width=width)

    def _on_mousewheel(self, event):
        self.canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        return "break"

    def _bind_mousewheel_recursive(self, widget):
        """递归绑定鼠标滚轮到所有子控件，确保在按钮上也能滚动。"""
        widget.bind("<MouseWheel>", self._on_mousewheel)
        for child in widget.winfo_children():
            self._bind_mousewheel_recursive(child)

    def clear(self):
        for child in self.inner.winfo_children():
            child.destroy()

    def add_button(self, row, col, text, command, width=16, font=("楷体", 10)):
        btn = tk.Button(
            self.inner,
            text=text,
            width=width,
            command=command,
            font=font,
            bg="#f8f9fa",
            fg="black",
            activebackground="#e2e6ea",
            relief="raised",
            bd=2,
        )
        btn.grid(row=row, column=col, padx=4, pady=4, sticky="nsew")
        # 确保鼠标在按钮上也能滚轮滚动
        btn.bind("<MouseWheel>", self._on_mousewheel)
        return btn

    def fill(self, items, callback, btn_width=16, btn_font=("楷体", 11, "bold")):
        """items: [(name, count, tag), ...]; callback(tag)"""
        self.clear()
        for i, (name, count, tag) in enumerate(items):
            r, c = divmod(i, self.cols)
            text = f"{name}（{count}家）"
            self.add_button(r, c, text, lambda t=tag: callback(t), width=btn_width, font=btn_font)
        for c in range(self.cols):
            self.inner.columnconfigure(c, weight=1)
        # 重新绑定所有子控件的滚轮事件
        self._bind_mousewheel_recursive(self.inner)



class ProvinceMapCanvas(tk.Canvas):
    """中国省份地图 Canvas（B 方案）：基于 GeoJSON 真实边界，支持悬停高亮和点击。"""

    # Excel 全名/简称 -> GeoJSON 中的短名
    NAME_ALIASES = {
        "北京": "北京", "北京市": "北京",
        "天津": "天津", "天津市": "天津",
        "上海": "上海", "上海市": "上海",
        "重庆": "重庆", "重庆市": "重庆",
        "黑龙江": "黑龙江", "黑龙江省": "黑龙江",
        "吉林": "吉林", "吉林省": "吉林",
        "辽宁": "辽宁", "辽宁省": "辽宁",
        "河北": "河北", "河北省": "河北",
        "山西": "山西", "山西省": "山西",
        "内蒙古": "内蒙古", "内蒙古自治区": "内蒙古",
        "陕西": "陕西", "陕西省": "陕西",
        "甘肃": "甘肃", "甘肃省": "甘肃",
        "青海": "青海", "青海省": "青海",
        "宁夏": "宁夏", "宁夏回族自治区": "宁夏",
        "新疆": "新疆", "新疆维吾尔自治区": "新疆",
        "山东": "山东", "山东省": "山东",
        "江苏": "江苏", "江苏省": "江苏",
        "安徽": "安徽", "安徽省": "安徽",
        "浙江": "浙江", "浙江省": "浙江",
        "福建": "福建", "福建省": "福建",
        "江西": "江西", "江西省": "江西",
        "河南": "河南", "河南省": "河南",
        "湖北": "湖北", "湖北省": "湖北",
        "湖南": "湖南", "湖南省": "湖南",
        "广东": "广东", "广东省": "广东",
        "广西": "广西", "广西壮族自治区": "广西",
        "海南": "海南", "海南省": "海南",
        "四川": "四川", "四川省": "四川",
        "贵州": "贵州", "贵州省": "贵州",
        "云南": "云南", "云南省": "云南",
        "西藏": "西藏", "西藏自治区": "西藏",
        "台湾": "台湾", "台湾省": "台湾",
        "香港": "香港", "香港特别行政区": "香港",
        "澳门": "澳门", "澳门特别行政区": "澳门",
    }

    def __init__(self, parent, province_counts, on_select, bg="#f8f9fa", **kwargs):
        super().__init__(parent, bg=bg, highlightthickness=0, **kwargs)
        self.province_counts = province_counts or {}
        self.on_select = on_select
        self.padding = 20

        self.features = []       # [{"name": short_name, "polygons": [[(lon,lat),...],...]}, ...]
        self.items = {}          # short_name -> {"polygons": [ids], "text": id, ...}
        self.hovered = None
        self.tooltip = None
        self._pending_draw = None

        self._load_geojson()
        self.bind("<Configure>", self._on_configure)
        self.bind("<Motion>", self._on_mouse_move)
        self.bind("<Button-1>", self._on_click)
        self.bind("<Leave>", self._on_leave)

        self._draw_map()

    def _load_geojson(self):
        path = resource_path("china_provinces.json")
        try:
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
        except Exception as e:
            print(f"加载地图数据失败: {e}")
            data = {"features": []}

        all_lons = []
        all_lats = []
        for feat in data.get("features", []):
            short_name = feat["properties"]["name"]
            geom = feat["geometry"]
            polygons = []
            if geom["type"] == "Polygon":
                polygons = [ring[:-1] for ring in geom["coordinates"]]
            elif geom["type"] == "MultiPolygon":
                for poly in geom["coordinates"]:
                    polygons.extend([ring[:-1] for ring in poly])
            if not polygons:
                continue
            for ring in polygons:
                for lon, lat in ring:
                    all_lons.append(lon)
                    all_lats.append(lat)
            self.features.append({"name": short_name, "polygons": polygons})

        if all_lons:
            self.min_lon = min(all_lons)
            self.max_lon = max(all_lons)
            self.min_lat = min(all_lats)
            self.max_lat = max(all_lats)
        else:
            self.min_lon, self.max_lon = 70, 140
            self.min_lat, self.max_lat = 15, 55

    def update_counts(self, province_counts):
        self.province_counts = province_counts or {}
        self._draw_map()

    def _normalize_name(self, name):
        if not name:
            return name
        return self.NAME_ALIASES.get(name.strip(), name.strip())

    def _color_for(self, count):
        if count <= 0:
            return "#cfe2ff"  # 无发债主体：淡蓝色
        counts = list(self.province_counts.values()) if self.province_counts else [1]
        max_count = max(counts) if counts else 1
        ratio = min(1.0, count / max(1, max_count))
        # 浅粉色 #ffd6e0 到深红色 #dc143c
        r = int(255 - ratio * (255 - 220))
        g = int(214 - ratio * (214 - 20))
        b = int(224 - ratio * (224 - 60))
        return f"#{r:02x}{g:02x}{b:02x}"

    def _lighten(self, color):
        return "#ff8fa3"

    def _project(self, lon, lat):
        width = max(1, self.winfo_width())
        height = max(1, self.winfo_height())
        pad = self.padding
        usable_w = width - 2 * pad
        usable_h = height - 2 * pad
        lon_range = self.max_lon - self.min_lon or 1
        lat_range = self.max_lat - self.min_lat or 1
        # 保持地图纵横比，居中显示
        map_ratio = lon_range / lat_range
        canvas_ratio = usable_w / usable_h
        if canvas_ratio > map_ratio:
            scale = usable_h / lat_range
            offset_x = (usable_w - lon_range * scale) / 2
            offset_y = 0
        else:
            scale = usable_w / lon_range
            offset_x = 0
            offset_y = (usable_h - lat_range * scale) / 2
        x = pad + offset_x + (lon - self.min_lon) * scale
        y = pad + offset_y + (self.max_lat - lat) * scale
        return x, y

    def _draw_map(self):
        self.delete("all")
        self.items = {}
        counts_by_short = {
            self._normalize_name(k): v
            for k, v in self.province_counts.items()
        }

        for feat in self.features:
            short_name = feat["name"]
            count = counts_by_short.get(short_name, 0)
            color = self._color_for(count)
            poly_ids = []
            centroid_x, centroid_y, total_area = 0, 0, 0
            for ring in feat["polygons"]:
                pts = []
                for lon, lat in ring:
                    x, y = self._project(lon, lat)
                    pts.extend([x, y])
                if len(pts) >= 6:
                    pid = self.create_polygon(pts, fill=color, outline="#ffffff", width=1, smooth=False)
                    poly_ids.append(pid)
                    n = len(pts) // 2
                    cx = sum(pts[i] for i in range(0, len(pts), 2)) / n
                    cy = sum(pts[i + 1] for i in range(0, len(pts), 2)) / n
                    area = abs(sum(
                        pts[i] * pts[(i + 3) % len(pts)] - pts[i + 1] * pts[(i + 2) % len(pts)]
                        for i in range(0, len(pts), 2)
                    )) / 2
                    centroid_x += cx * area
                    centroid_y += cy * area
                    total_area += area

            # 计算省份边界框面积，据此决定是否显示文字
            all_x = [pts[i] for i in range(0, len(pts), 2)]
            all_y = [pts[i + 1] for i in range(0, len(pts), 2)]
            bbox_w = max(all_x) - min(all_x)
            bbox_h = max(all_y) - min(all_y)
            bbox_area = bbox_w * bbox_h

            if total_area > 0:
                centroid_x /= total_area
                centroid_y /= total_area
            else:
                first = feat["polygons"][0][0]
                centroid_x, centroid_y = self._project(first[0], first[1])

            tid = None
            if bbox_area >= 600:
                display_name = short_name
                name_font = ("Microsoft YaHei", 12, "bold")
                if bbox_area < 1800:
                    # 较小省份只显示名称，字体稍小
                    name_font = ("Microsoft YaHei", 10, "bold")
                text = display_name
                if bbox_area >= 1800 and count > 0:
                    text = f"{display_name}（{count}家）"
                tid = self.create_text(
                    centroid_x, centroid_y,
                    text=text, font=name_font,
                    fill="black", justify="center"
                )

            self.items[short_name] = {
                "polygons": poly_ids,
                "text": tid,
                "base_color": color,
            }

    def _on_configure(self, event):
        # 立即重绘；连续 resize 时通过 cancel/after 做少量防抖
        self._draw_map()
        if self._pending_draw:
            self.after_cancel(self._pending_draw)
        self._pending_draw = self.after(50, self._draw_map)

    def _point_in_polygons(self, lon, lat):
        """射线法判断经纬度点落在哪个省份。"""
        for feat in self.features:
            for ring in feat["polygons"]:
                inside = False
                n = len(ring)
                j = n - 1
                for i in range(n):
                    xi, yi = ring[i]
                    xj, yj = ring[j]
                    if ((yi > lat) != (yj > lat)) and (lon < (xj - xi) * (lat - yi) / (yj - yi + 1e-12) + xi):
                        inside = not inside
                    j = i
                if inside:
                    return feat["name"]
        return None

    def _hit_test(self, x, y):
        width = max(1, self.winfo_width())
        height = max(1, self.winfo_height())
        pad = self.padding
        usable_w = width - 2 * pad
        usable_h = height - 2 * pad
        lon_range = self.max_lon - self.min_lon or 1
        lat_range = self.max_lat - self.min_lat or 1
        map_ratio = lon_range / lat_range
        canvas_ratio = usable_w / usable_h
        if canvas_ratio > map_ratio:
            scale = usable_h / lat_range
            offset_x = (usable_w - lon_range * scale) / 2
            offset_y = 0
        else:
            scale = usable_w / lon_range
            offset_x = 0
            offset_y = (usable_h - lat_range * scale) / 2
        lon = self.min_lon + (x - pad - offset_x) / scale
        lat = self.max_lat - (y - pad - offset_y) / scale
        return self._point_in_polygons(lon, lat)

    def _count_for(self, short_name):
        count = self.province_counts.get(short_name, 0)
        for k, v in self.province_counts.items():
            if self._normalize_name(k) == short_name:
                count = v
                break
        return count

    def _on_mouse_move(self, event):
        province = self._hit_test(event.x, event.y)
        if province == self.hovered:
            self._move_tooltip(event.x, event.y)
            return
        if self.hovered and self.hovered in self.items:
            item = self.items[self.hovered]
            for pid in item["polygons"]:
                self.itemconfig(pid, fill=item["base_color"])
        self.hovered = province
        if province:
            item = self.items[province]
            for pid in item["polygons"]:
                self.itemconfig(pid, fill=self._lighten(item["base_color"]))
            self._show_tooltip(event.x, event.y, province, self._count_for(province))
        else:
            self._hide_tooltip()

    def _on_click(self, event):
        province = self._hit_test(event.x, event.y)
        if province and self.on_select:
            raw_name = province
            for k in self.province_counts:
                if self._normalize_name(k) == province:
                    raw_name = k
                    break
            self.on_select(raw_name)

    def _on_leave(self, event):
        if self.hovered and self.hovered in self.items:
            item = self.items[self.hovered]
            for pid in item["polygons"]:
                self.itemconfig(pid, fill=item["base_color"])
        self.hovered = None
        self._hide_tooltip()

    def _show_tooltip(self, x, y, province, count):
        if self.tooltip is None:
            self.tooltip = tk.Toplevel(self)
            self.tooltip.wm_overrideredirect(True)
            self.tooltip_label = ttk.Label(
                self.tooltip,
                text="",
                font=("Microsoft YaHei", 10, "bold"),
                background="#333333",
                foreground="white",
                padding=(8, 4),
            )
            self.tooltip_label.pack()
        self.tooltip_label.config(text=f"{province}\n{count} 家")
        self._move_tooltip(x, y)
        self.tooltip.deiconify()

    def _move_tooltip(self, x, y):
        if self.tooltip:
            cx = self.winfo_rootx() + x + 15
            cy = self.winfo_rooty() + y + 15
            self.tooltip.wm_geometry(f"+{cx}+{cy}")

    def _hide_tooltip(self):
        if self.tooltip:
            self.tooltip.withdraw()


class BondBrowserApp(tk.Tk):
    EXCEL_FILE = "债券数据浏览器20260616-v6.xlsx"
    USERS_FILE = "债市读者的问答社区-用户活跃-成员数据报表（所有成员).xlsx"

    # 兼容引用（也保留类属性，方便内部统一使用 self.XXX）
    COL_BOND_TYPE = COL_BOND_TYPE
    COL_ISSUER = COL_ISSUER
    COL_PROVINCE = COL_PROVINCE
    COL_CITY = COL_CITY
    COL_DISTRICT = COL_DISTRICT
    COL_INDUSTRY = COL_INDUSTRY
    COL_ENTERPRISE_NATURE = COL_ENTERPRISE_NATURE
    COL_CENTRAL = COL_CENTRAL
    COL_MUNICIPAL = COL_MUNICIPAL
    COL_PROVINCIAL = COL_PROVINCIAL
    COL_COUNTY = COL_COUNTY
    COL_CONTROLLER = COL_CONTROLLER
    COL_SHAREHOLDER = COL_SHAREHOLDER
    COL_LGFV_SUBSIDIARY = COL_LGFV_SUBSIDIARY
    COL_LGFV_LEVEL = COL_LGFV_LEVEL
    COL_BOND_COUNT = COL_BOND_COUNT
    COL_TOTAL_ASSETS = COL_TOTAL_ASSETS
    COL_TOTAL_EQUITY = COL_TOTAL_EQUITY
    COL_TOTAL_REVENUE = COL_TOTAL_REVENUE
    COL_NET_PROFIT = COL_NET_PROFIT
    COL_CASH_FLOW_2024 = COL_CASH_FLOW_2024
    COL_CASH_FLOW_2025 = COL_CASH_FLOW_2025
    COL_ROA = COL_ROA
    COL_REVENUE_RATIO = COL_REVENUE_RATIO
    COL_REVENUE = COL_REVENUE
    COL_PROJECT_NAME = COL_PROJECT_NAME
    COL_BOND_SHORT = COL_BOND_SHORT
    COL_BOND_CODE = COL_BOND_CODE
    COL_BOND_FULL = COL_BOND_FULL
    COL_ISSUE_AMOUNT = COL_ISSUE_AMOUNT
    COL_BOND_TERM = COL_BOND_TERM
    COL_START_DATE = COL_START_DATE
    COL_COUPON_RATE = COL_COUPON_RATE
    COL_USE_OF_PROCEEDS = COL_USE_OF_PROCEEDS
    COL_BOND_RATING = COL_BOND_RATING
    COL_ISSUER_RATING = COL_ISSUER_RATING
    COL_GUARANTOR = COL_GUARANTOR
    COL_GUARANTOR_RATING = COL_GUARANTOR_RATING

    MUNICIPALITIES = {"北京市", "上海市", "天津市", "重庆市"}

    INSERT_BATCH = 100

    # 各分类页显示的列：("显示名称", 原始列索引 或 lambda row: 值)
    TAB_COLUMNS = {
        "汇总": [
            ("所属区域", lambda r: " ".join(filter(None, [r[COL_PROVINCE], r[COL_CITY], r[COL_DISTRICT]]))),
            ("发行人名称", COL_ISSUER),
            ("主体评级", COL_ISSUER_RATING),
            ("债券名称", COL_BOND_SHORT),
            ("所属行业", COL_INDUSTRY),
            ("实控人", COL_CONTROLLER),
            ("首次发债日期", COL_START_DATE),
        ],
        "主体信息": [
            ("所属区域", lambda r: " ".join(filter(None, [r[COL_PROVINCE], r[COL_CITY], r[COL_DISTRICT]]))),
            ("发行人名称", COL_ISSUER),
            ("所属省", COL_PROVINCE),
            ("所属市", COL_CITY),
            ("所属区/县", COL_DISTRICT),
            ("申万一级行业", COL_INDUSTRY),
            ("企业性质", COL_ENTERPRISE_NATURE),
            ("实际控制人", COL_CONTROLLER),
            ("股东", COL_SHAREHOLDER),
            ("是否城投子公司", COL_LGFV_SUBSIDIARY),
            ("债券存量只数", COL_BOND_COUNT),
        ],
        "债券信息": [
            ("发行人名称", COL_ISSUER),
            ("债券简称", COL_BOND_SHORT),
            ("债券代码", COL_BOND_CODE),
            ("债券全称", COL_BOND_FULL),
            ("发行总额（亿元）", COL_ISSUE_AMOUNT),
            ("债券期限（年）", COL_BOND_TERM),
            ("起息日期", COL_START_DATE),
            ("发行票面利率（%）", COL_COUPON_RATE),
            ("募集资金用途", COL_USE_OF_PROCEEDS),
            ("发行时债券评级", COL_BOND_RATING),
            ("担保人名称", COL_GUARANTOR),
            ("最新担保人评级", COL_GUARANTOR_RATING),
        ],
        "财务信息": [
            ("发行人名称", COL_ISSUER),
            ("2025年末资产总计（亿元）", COL_TOTAL_ASSETS),
            ("2025年末所有者权益总计（亿元）", COL_TOTAL_EQUITY),
            ("2025年营业总收入（亿元）", COL_TOTAL_REVENUE),
            ("2025年净利润（亿元）", COL_NET_PROFIT),
            ("2024年度经营活动现金流净额", COL_CASH_FLOW_2024),
            ("2025年度经营活动现金流净额", COL_CASH_FLOW_2025),
            ("2025年总资产报酬率(%)", COL_ROA),
        ],
        "第一大收入构成": [
            ("发行人名称", COL_ISSUER),
            ("营业收入占比(%)", COL_REVENUE_RATIO),
            ("营业收入（亿元）", COL_REVENUE),
            ("项目名称", COL_PROJECT_NAME),
        ],
    }

    def __init__(self):
        super().__init__()
        self.title("债券数据浏览器 V3")
        self.geometry("1200x700")
        self.minsize(900, 500)
        self.resizable(True, True)

        # 全局样式：滚动条加粗，标题类字体加粗
        self._setup_styles()

        self.protocol("WM_DELETE_WINDOW", self._on_close)
        self.bind("<Escape>", lambda e: self._on_close())
        self.bind("<Alt-F4>", lambda e: self._on_close())

        self.headers = []
        self.rows = []              # 全部数据
        self.users = {}             # 用户昵称 -> 到期时间
        self.authenticated = False

        self.current_bond_type = None
        self.current_level = 0      # 0=债券类型, 1=全国, 2=省份, 3=城市
        self.current_province = None
        self.current_city = None

        self._build_ui()
        self.after(100, self._load_data)

    # ------------------------------------------------------------------
    # UI 构建
    # ------------------------------------------------------------------
    def _build_ui(self):
        self.top_frame = ttk.Frame(self, padding=10)
        self.top_frame.pack(fill=tk.X)
        self.top_frame.columnconfigure(2, weight=1)

        self.btn_back = ttk.Button(self.top_frame, text="← 返回", command=self._go_back)
        self.btn_back.grid(row=0, column=0, padx=(0, 8))

        self.btn_home = ttk.Button(self.top_frame, text="首页", command=self._go_home)
        self.btn_home.grid(row=0, column=1, padx=(0, 16))

        self.title_lbl = ttk.Label(
            self.top_frame,
            text="正在加载数据...",
            font=("Microsoft YaHei", 14, "bold"),
        )
        self.title_lbl.grid(row=0, column=2, padx=(0, 16), sticky="w")

        self.status_lbl = ttk.Label(self.top_frame, text="")
        self.status_lbl.grid(row=0, column=3, padx=(0, 12))

        self.btn_exit = ttk.Button(self.top_frame, text="退出", command=self._on_close)
        self.btn_exit.grid(row=0, column=4)

        self.content_frame = ttk.Frame(self)
        self.content_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0, 10))
        self.content_frame.rowconfigure(0, weight=1)
        self.content_frame.columnconfigure(0, weight=1)

        self.nav_frame = ttk.Frame(self.content_frame)
        self.nav_frame.grid(row=0, column=0, sticky="nsew")
        self.nav_frame.rowconfigure(0, weight=1)
        self.nav_frame.columnconfigure(0, weight=1)

        # 债券类型选择专用居中面板
        self.bond_type_frame = ttk.Frame(self.nav_frame)
        self.bond_type_frame.bind("<Configure>", lambda e: self._center_bond_type_frame())
        self.nav_frame.bind("<Configure>", lambda e: self._center_bond_type_frame())

        ttk.Label(
            self.bond_type_frame,
            text="请选择债券类型",
            font=("楷体", 20, "bold"),
        ).pack(pady=(0, 30))

        self.bond_type_btn_frame = ttk.Frame(self.bond_type_frame)
        self.bond_type_btn_frame.pack()

        # 导航网格（全国/省份/城市）
        self.nav_sub_frame = ttk.Frame(self.nav_frame)
        self.nav_sub_frame.grid(row=0, column=0, sticky="nsew")
        self.nav_sub_frame.rowconfigure(2, weight=1)
        self.nav_sub_frame.columnconfigure(0, weight=1)
        self.nav_sub_frame.columnconfigure(1, weight=4)
        self.nav_sub_frame.grid_remove()

        self.nav_hint = ttk.Label(self.nav_sub_frame, text="")
        self.nav_hint.grid(row=0, column=0, columnspan=2, sticky="ew", pady=(0, 4))

        self.central_btn = tk.Button(
            self.nav_sub_frame,
            text="央企（0家）",
            font=("Microsoft YaHei", 12, "bold"),
            bg="#6f42c1",
            fg="white",
            activebackground="#5a32a3",
            width=20,
            command=self._on_central_btn_click,
        )
        self.central_btn.grid(row=1, column=0, columnspan=2, pady=(0, 10))
        self.central_btn.grid_remove()

        # 左侧省份按钮列表（含央企）
        self.left_frame = ttk.Frame(self.nav_sub_frame)
        self.left_frame.rowconfigure(0, weight=1)
        self.left_frame.columnconfigure(0, weight=1)

        self.province_list = ScrollableButtonGrid(self.left_frame, cols=1)
        self.province_list.grid(row=0, column=0, sticky="nsew")

        # 右侧地图
        self.right_frame = ttk.Frame(self.nav_sub_frame)
        self.right_frame.rowconfigure(0, weight=1)
        self.right_frame.columnconfigure(0, weight=1)

        self.province_map = ProvinceMapCanvas(
            self.right_frame,
            province_counts={},
            on_select=None,
        )
        self.province_map.grid(row=0, column=0, sticky="nsew")
        self.province_map.grid_remove()

        # 省/市/区县层级全宽按钮网格
        self.button_grid = ScrollableButtonGrid(self.nav_sub_frame, cols=5)

        self.detail_frame = ttk.Frame(self.content_frame)
        self.detail_frame.grid(row=0, column=0, sticky="nsew")
        self.detail_frame.rowconfigure(2, weight=1)
        self.detail_frame.columnconfigure(0, weight=1)
        self._build_detail_view(self.detail_frame)

        self.detail_frame.grid_remove()

    def _center_bond_type_frame(self):
        """让债券类型选择面板在窗口中居中；只在债券类型选择层生效。"""
        if self.current_level != 0:
            return
        self.nav_frame.update_idletasks()
        w = self.nav_frame.winfo_width()
        h = self.nav_frame.winfo_height()
        fw = self.bond_type_frame.winfo_width()
        fh = self.bond_type_frame.winfo_height()
        x = max(0, (w - fw) // 2)
        y = max(0, (h - fh) // 2)
        self.bond_type_frame.place(x=x, y=y)

    def _build_detail_view(self, parent):
        self.detail_title = ttk.Label(parent, text="", font=("Microsoft YaHei", 12, "bold"))
        self.detail_title.grid(row=0, column=0, sticky="ew", pady=(0, 5))

        # 自定义 Tab 按钮栏
        self.tab_bar = ttk.Frame(parent)
        self.tab_bar.grid(row=1, column=0, sticky="ew", pady=(0, 5))

        self.tab_trees = {}        # tab_name -> (frozen_tree, scroll_tree)
        self.tab_frames = {}
        self.tab_buttons = {}
        self.current_tab = None

        tab_names = ["汇总", "主体信息", "债券信息", "财务信息", "第一大收入构成"]
        for idx, tab_name in enumerate(tab_names):
            frame = tk.Frame(parent, relief="groove", bd=2, bg="#f8f9fa")
            frame.grid(row=2, column=0, sticky="nsew", pady=(0, 5))
            self.tab_frames[tab_name] = frame

            tree_frame = ttk.Frame(frame)
            tree_frame.pack(fill=tk.BOTH, expand=True, padx=4, pady=4)
            tree_frame.rowconfigure(0, weight=1)
            tree_frame.columnconfigure(1, weight=1)

            # 冻结列（发行人名称）放在左侧
            frozen_frame = ttk.Frame(tree_frame)
            frozen_frame.grid(row=0, column=0, sticky="nsew")
            frozen_frame.rowconfigure(0, weight=1)
            frozen_frame.columnconfigure(0, weight=1)

            frozen_tree = ttk.Treeview(
                frozen_frame,
                show="headings",
                selectmode="browse",
                height=10,
            )
            frozen_tree.grid(row=0, column=0, sticky="nsew")

            # 可滚动列放在右侧
            right_frame = ttk.Frame(tree_frame)
            right_frame.grid(row=0, column=1, sticky="nsew")
            right_frame.rowconfigure(0, weight=1)
            right_frame.columnconfigure(0, weight=1)

            vsb = ttk.Scrollbar(right_frame, orient=tk.VERTICAL)
            vsb.grid(row=0, column=1, sticky="ns")
            hsb = ttk.Scrollbar(tree_frame, orient=tk.HORIZONTAL)
            hsb.grid(row=1, column=0, columnspan=2, sticky="ew")

            scroll_tree = ttk.Treeview(
                right_frame,
                xscrollcommand=hsb.set,
                show="headings",
                selectmode="browse",
                height=10,
            )
            scroll_tree.grid(row=0, column=0, sticky="nsew")
            vsb.config(command=scroll_tree.yview)
            hsb.config(command=scroll_tree.xview)

            # 垂直滚动同步：scroll_tree 为主，frozen_tree 跟随
            def sync_yscroll(ft=frozen_tree, st=scroll_tree, v=vsb):
                def callback(first, last):
                    v.set(first, last)
                    ft.yview_moveto(first)
                return callback
            scroll_tree.config(yscrollcommand=sync_yscroll())

            # 鼠标滚轮：两个 tree 都绑定到 scroll_tree 的滚动
            def on_wheel(event, t=scroll_tree):
                t.yview_scroll(int(-1 * (event.delta / 120)), "units")
                return "break"
            scroll_tree.bind("<MouseWheel>", on_wheel)
            frozen_tree.bind("<MouseWheel>", on_wheel)
            scroll_tree.bind("<Shift-MouseWheel>", lambda e, t=scroll_tree: t.xview_scroll(int(-1 * (e.delta / 120)), "units") or "break")

            # 选择同步
            def sync_select(event, ft=frozen_tree, st=scroll_tree):
                src = event.widget
                sel = src.selection()
                if not sel:
                    return
                target = st if src == ft else ft
                target.selection_set(sel)
                target.see(sel[0])
            scroll_tree.bind("<<TreeviewSelect>>", sync_select)
            frozen_tree.bind("<<TreeviewSelect>>", sync_select)

            self._setup_copy_menu(scroll_tree)
            self._setup_copy_menu(frozen_tree)
            self.tab_trees[tab_name] = (frozen_tree, scroll_tree)

            btn = tk.Button(
                self.tab_bar,
                text=tab_name,
                font=("Microsoft YaHei", 10, "bold"),
                width=14,
                relief="raised",
                bd=2,
                bg="#f8f9fa",
                fg="#333333",
                activebackground="#e2e6ea",
                highlightbackground="#ced4da",
                highlightthickness=1,
                command=lambda t=tab_name: self._switch_tab(t),
            )
            btn.pack(side=tk.LEFT, padx=4, pady=4)
            self.tab_buttons[tab_name] = btn

            if idx < len(tab_names) - 1:
                ttk.Separator(self.tab_bar, orient=tk.VERTICAL).pack(side=tk.LEFT, fill=tk.Y, padx=2, pady=4)

        self._switch_tab("汇总")

    def _setup_copy_menu(self, tree):
        """为 Treeview 增加右键复制单元格功能。"""
        menu = tk.Menu(tree, tearoff=0)
        menu.add_command(label="复制", command=lambda: self._copy_cell(tree))

        def on_right_click(event):
            region = tree.identify("region", event.x, event.y)
            if region not in ("cell", "tree"):
                return "break"
            row_id = tree.identify_row(event.y)
            col_id = tree.identify_column(event.x)
            if row_id and col_id:
                tree.selection_set(row_id)
                tree.focus(row_id)
                tree._copy_row_id = row_id
                tree._copy_col_id = col_id
                menu.post(event.x_root, event.y_root)
            return "break"

        tree.bind("<Button-3>", on_right_click)

    def _copy_cell(self, tree):
        row_id = getattr(tree, "_copy_row_id", None)
        col_id = getattr(tree, "_copy_col_id", None)
        if not row_id or not col_id:
            return
        values = tree.item(row_id, "values")
        try:
            # Treeview identify_column 返回 #1、#2…表示第 1/2 个显示列
            idx = int(str(col_id).replace("#", "")) - 1
        except ValueError:
            return
        if 0 <= idx < len(values):
            self.clipboard_clear()
            self.clipboard_append(str(values[idx]))

    # ------------------------------------------------------------------
    # 数据加载
    # ------------------------------------------------------------------
    def _load_data(self):
        path = resource_path(self.EXCEL_FILE)
        if not os.path.exists(path):
            path = filedialog.askopenfilename(
                title="请选择债券数据 Excel 文件",
                filetypes=[("Excel 文件", "*.xlsx *.xls")],
            )
            if not path:
                messagebox.showerror("未找到数据", "未找到债券数据文件，程序将退出。")
                self.destroy()
                return

        try:
            wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
            ws = wb[wb.sheetnames[0]]
            all_rows = list(ws.iter_rows(values_only=True))
            wb.close()
        except Exception as e:
            messagebox.showerror("读取失败", f"读取债券数据失败：{e}")
            self.destroy()
            return

        if len(all_rows) < 2:
            messagebox.showerror("数据为空", "Excel 中没有有效数据。")
            self.destroy()
            return

        # v5 表头在第一行
        self.headers = [self._fmt_header(v) for v in all_rows[0]]
        self.rows = []
        for r in all_rows[1:]:
            row = tuple(self._fmt_value(v) for v in r)
            if len(row) < len(self.headers):
                row = row + ("",) * (len(self.headers) - len(row))
            elif len(row) > len(self.headers):
                row = row[: len(self.headers)]
            self.rows.append(row)

        self._load_users()
        self._setup_detail_columns()
        self._show_auth_popup()

    def _load_users(self):
        users_path = resource_path(self.USERS_FILE)
        self.users = {}
        if not os.path.exists(users_path):
            messagebox.showwarning(
                "未找到用户表",
                f"未找到用户数据文件：{self.USERS_FILE}\n将跳过认证检查。",
            )
            return
        try:
            wb = openpyxl.load_workbook(users_path, data_only=True, read_only=True)
            ws = wb[wb.sheetnames[0]]
            header = None
            nick_idx = None
            expiry_idx = None
            for row in ws.iter_rows(values_only=True):
                if header is None:
                    header = [str(c or "").strip() for c in row]
                    nick_idx = header.index("用户昵称")
                    expiry_idx = header.index("到期时间")
                    continue
                if not row or nick_idx >= len(row) or not row[nick_idx]:
                    continue
                nickname = str(row[nick_idx]).strip()
                expiry = row[expiry_idx] if expiry_idx < len(row) else None
                if isinstance(expiry, str):
                    expiry = self._parse_date(expiry)
                self.users[nickname] = expiry
            wb.close()
        except Exception as e:
            messagebox.showwarning("用户表读取异常", f"读取用户表失败：{e}\n将跳过认证检查。")

    @staticmethod
    def _parse_date(value):
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d", "%Y%m%d"):
            try:
                return datetime.strptime(value.strip(), fmt)
            except Exception:
                continue
        return None

    def _setup_styles(self):
        """配置全局 ttk 样式：滚动条加粗，标题字体加粗等。"""
        style = ttk.Style(self)
        # 滚动条加粗
        style.configure("Vertical.TScrollbar", width=34, background="#0d6efd", troughcolor="#e9ecef", bordercolor="#0d6efd", arrowcolor="white")
        style.configure("Horizontal.TScrollbar", width=34, background="#0d6efd", troughcolor="#e9ecef", bordercolor="#0d6efd", arrowcolor="white")
        # Treeview 表头加粗
        style.configure("Treeview.Heading", font=("Microsoft YaHei", 10, "bold"))

    def _setup_detail_columns(self):
        """仅初始化列标识与表头文字，宽度在 _show_detail 中按内容自适应。
        发行人名称列默认冻结在左侧。"""
        for tab_name, (frozen_tree, scroll_tree) in self.tab_trees.items():
            cols = self.TAB_COLUMNS[tab_name]
            # 找到要冻结的列：优先“发行人名称”，否则第一列
            freeze_idx = 0
            for i, (name, _) in enumerate(cols):
                if name == "发行人名称":
                    freeze_idx = i
                    break

            frozen_tree["columns"] = [f"{tab_name}_f{freeze_idx}"]
            frozen_tree.heading(f"{tab_name}_f{freeze_idx}", text=cols[freeze_idx][0])
            frozen_tree.column(f"{tab_name}_f{freeze_idx}", width=80, anchor="w", stretch=False)

            scroll_cols = [i for i in range(len(cols)) if i != freeze_idx]
            scroll_tree["columns"] = [f"{tab_name}_{i}" for i in scroll_cols]
            for i in scroll_cols:
                cid = f"{tab_name}_{i}"
                scroll_tree.heading(cid, text=cols[i][0])
                scroll_tree.column(cid, width=80, anchor="w", stretch=False)

            self.tab_trees[tab_name] = (frozen_tree, scroll_tree, freeze_idx)

    def _auto_fit_columns(self, rows):
        """根据表头和数据内容自动调整所有 Tab 的列宽（含冻结列）。"""
        for tab_name, (frozen_tree, scroll_tree, freeze_idx) in self.tab_trees.items():
            cols = self.TAB_COLUMNS[tab_name]
            widths = [len(display_name) for display_name, _ in cols]

            for r in rows:
                for i, (_, src) in enumerate(cols):
                    if callable(src):
                        try:
                            val = str(src(r))
                        except Exception:
                            val = ""
                    else:
                        val = str(r[src] if src < len(r) else "")
                    widths[i] = max(widths[i], len(val))

            # 冻结列
            w = max(80, min(400, widths[freeze_idx] * 12 + 30))
            frozen_tree.column(f"{tab_name}_f{freeze_idx}", width=w)

            # 滚动列
            for i in range(len(cols)):
                if i == freeze_idx:
                    continue
                cid = f"{tab_name}_{i}"
                w = max(80, min(600, widths[i] * 12 + 30))
                scroll_tree.column(cid, width=w)

    # ------------------------------------------------------------------
    # 认证弹窗
    # ------------------------------------------------------------------
    def _show_auth_popup(self):
        if self.authenticated:
            self._show_bond_type_selection()
            return

        popup = tk.Toplevel(self)
        popup.title("身份验证")
        popup.geometry("360x160")
        popup.resizable(False, False)
        popup.transient(self)
        popup.grab_set()
        popup.protocol("WM_DELETE_WINDOW", popup.destroy)

        ttk.Label(popup, text="请输入用户星球昵称：", font=("Microsoft YaHei", 10)).pack(pady=(15, 5))

        entry = ttk.Entry(popup, width=30)
        entry.pack(pady=5)
        entry.focus_set()

        result_lbl = ttk.Label(popup, text="", foreground="red")
        result_lbl.pack(pady=(0, 5))

        def on_confirm():
            nickname = entry.get().strip()
            ok, msg = self._check_auth(nickname)
            if ok:
                self.authenticated = True
                popup.destroy()
                self._show_bond_type_selection()
            else:
                result_lbl.config(text=msg)

        def on_cancel():
            popup.destroy()

        btn_frame = ttk.Frame(popup)
        btn_frame.pack(pady=5)
        ttk.Button(btn_frame, text="确认", command=on_confirm, width=10).pack(side=tk.LEFT, padx=8)
        ttk.Button(btn_frame, text="取消", command=on_cancel, width=10).pack(side=tk.LEFT, padx=8)

        entry.bind("<Return>", lambda e: on_confirm())

    def _check_auth(self, nickname):
        if not nickname:
            return False, "请输入星球昵称"
        if not self.users:
            return True, ""
        expiry = self.users.get(nickname)
        if expiry is None:
            return False, "请输入正确的星球昵称"
        if isinstance(expiry, datetime) and expiry < datetime.now():
            return False, "星球昵称正确，用户已到期"
        return True, ""

    # ------------------------------------------------------------------
    # 债券类型选择
    # ------------------------------------------------------------------
    def _show_bond_type_selection(self):
        self.current_level = 0
        self.current_bond_type = None
        self.detail_frame.grid_remove()
        self.nav_sub_frame.grid_remove()
        self.nav_frame.grid()
        self.bond_type_frame.place(x=0, y=0)
        self.bond_type_frame.lift()
        self._center_bond_type_frame()

        # 清除旧按钮
        for child in self.bond_type_btn_frame.winfo_children():
            child.destroy()

        counts = {}
        for r in self.rows:
            bt = r[self.COL_BOND_TYPE]
            if bt:
                counts[bt] = counts.get(bt, 0) + 1

        for bt in ["公司债券", "协会债券"]:
            if bt not in counts:
                continue
            btn = tk.Button(
                self.bond_type_btn_frame,
                text=f"{bt}\n({counts[bt]}家)",
                font=("楷体", 16),
                width=16,
                height=4,
                relief="raised",
                bd=2,
                command=lambda t=bt: self._on_bond_type_click(t),
            )
            btn.pack(side=tk.LEFT, padx=25, pady=10)

        total = sum(counts.values())
        self.title_lbl.config(text="债券数据浏览器 V3")
        self.status_lbl.config(text=f"共 {total} 条债券数据")

    def _on_bond_type_click(self, bond_type):
        self.current_bond_type = bond_type
        self.bond_type_frame.place_forget()
        self.nav_sub_frame.grid()
        self._show_level(1)

    # ------------------------------------------------------------------
    # 企业层级判断
    # ------------------------------------------------------------------
    def _is_municipality(self, province):
        return province in self.MUNICIPALITIES

    def _row_level(self, r, province):
        """判断一行数据在该省份下属于哪个层级。央企已在国家级单独处理。"""
        in_muni = self._is_municipality(province)
        if r[self.COL_PROVINCIAL] == "是":
            return "provincial"
        if r[self.COL_MUNICIPAL] == "是":
            return "municipal"
        if r[self.COL_COUNTY] == "是":
            return "county"
        # 无明确层级标记时，按所属市/区县字段兜底判断
        city = r[self.COL_CITY]
        district = r[self.COL_DISTRICT]
        if not city and not district:
            # 直辖市无市/区县字段时，视为市本级；其余省份视为省本级
            return "municipal" if in_muni else "provincial"
        if city and not district:
            return "municipal"
        return "county"

    def _filter_rows(self, bond_type=None, province=None, city=None, district=None, exclude_central=True):
        result = []
        for r in self.rows:
            if bond_type is not None and r[self.COL_BOND_TYPE] != bond_type:
                continue
            if exclude_central and r[self.COL_CENTRAL] == "是":
                continue
            if province is not None and r[self.COL_PROVINCE] != province:
                continue
            if city is not None and r[self.COL_CITY] != city:
                continue
            if district is not None and r[self.COL_DISTRICT] != district:
                continue
            result.append(r)
        return result

    def _central_rows(self, bond_type):
        return [r for r in self.rows if r[self.COL_BOND_TYPE] == bond_type and r[self.COL_CENTRAL] == "是"]

    # ------------------------------------------------------------------
    # 导航逻辑
    # ------------------------------------------------------------------
    def _on_central_btn_click(self):
        rows = self._central_rows(self.current_bond_type)
        if rows:
            self._show_detail("— 央企", rows)
        else:
            messagebox.showinfo("提示", "当前债券类型下没有央企数据。")

    def _show_level(self, level, province=None, city=None):
        self.current_level = level
        self.current_province = province
        self.current_city = city

        self.detail_frame.grid_remove()
        self.bond_type_frame.place_forget()
        self.nav_frame.grid()
        self.nav_sub_frame.grid()

        bt = self.current_bond_type

        if level == 1:
            self.title_lbl.config(text=f"{bt} — 全国省份/央企")
            self.nav_hint.config(text="左侧为省份/央企按钮，右侧为地图；点击任意省份进入该省下辖层级。")
            items = self._get_national_items(bt)
            central = next((c for n, c, t in items if t == "CENTRAL"), 0)
            province_counts = {p: c for p, c, t in items if t not in ("CENTRAL",)}
            self.central_btn.config(text=f"央企（{central}家）")
            self.central_btn.grid()
            self.button_grid.grid_remove()
            self.left_frame.grid(row=2, column=0, sticky="nsew", padx=(0, 8))
            self.right_frame.grid(row=2, column=1, sticky="nsew")
            self.province_list.fill(items, self._on_national_item_click, btn_width=8)
            self.province_map.update_counts(province_counts)
            self.province_map.on_select = lambda name: self._show_level(2, province=name)
            self.province_map.grid()
            self.status_lbl.config(text=self._national_status(bt, items))

        else:
            self.central_btn.grid_remove()
            self.left_frame.grid_remove()
            self.right_frame.grid_remove()
            self.province_map.grid_remove()
            self.button_grid.grid(row=2, column=0, columnspan=2, sticky="nsew")

            if level == 2:
                self.title_lbl.config(text=f"{bt} — {province} 下辖层级")
                self.nav_hint.config(text="点击层级查看发债主体明细。")
                items = self._get_province_items(province)
                self.button_grid.fill(items, self._on_province_item_click, btn_width=10)
                self.status_lbl.config(text=self._province_status(province, items))

            elif level == 3:
                self.title_lbl.config(text=f"{bt} — {province} · {city} 下辖层级")
                self.nav_hint.config(text="点击层级查看发债主体明细。")
                items = self._get_city_items(province, city)
                self.button_grid.fill(items, self._on_city_item_click, btn_width=10)
                self.status_lbl.config(text=self._city_status(items))

    def _get_national_items(self, bond_type):
        items = []
        central = self._central_rows(bond_type)
        if central:
            items.append(("央企", len(central), "CENTRAL"))

        province_counts = {}
        for r in self._filter_rows(bond_type=bond_type, exclude_central=True):
            p = r[self.COL_PROVINCE]
            if p:
                province_counts[p] = province_counts.get(p, 0) + 1
        for p, c in sorted(province_counts.items(), key=lambda x: -x[1]):
            items.append((p, c, p))
        return items

    def _get_province_items(self, province):
        items = []
        in_muni = self._is_municipality(province)

        if in_muni:
            muni_count = 0
            district_counts = {}
            for r in self._filter_rows(bond_type=self.current_bond_type, province=province):
                level = self._row_level(r, province)
                if level in ("provincial", "municipal"):
                    muni_count += 1
                elif level == "county":
                    d = r[self.COL_DISTRICT]
                    if d:
                        district_counts[d] = district_counts.get(d, 0) + 1
            items.append(("市本级", muni_count, "MUNICIPAL"))
            for d, c in sorted(district_counts.items(), key=lambda x: -x[1]):
                items.append((d, c, d))
        else:
            prov_count = 0
            city_counts = {}
            for r in self._filter_rows(bond_type=self.current_bond_type, province=province):
                level = self._row_level(r, province)
                if level == "provincial":
                    prov_count += 1
                elif level in ("municipal", "county"):
                    c = r[self.COL_CITY]
                    if c:
                        city_counts[c] = city_counts.get(c, 0) + 1
            items.append(("省本级", prov_count, "PROVINCIAL"))
            for c, count in sorted(city_counts.items(), key=lambda x: -x[1]):
                items.append((c, count, c))
        return items

    def _get_city_items(self, province, city):
        items = []
        muni_count = 0
        district_counts = {}
        in_muni = self._is_municipality(province)
        for r in self._filter_rows(bond_type=self.current_bond_type, province=province, city=city):
            level = self._row_level(r, province)
            if in_muni:
                # 直辖市：省级/市级都归入市本级
                if level in ("provincial", "municipal"):
                    muni_count += 1
                elif level == "county":
                    d = r[self.COL_DISTRICT]
                    if d:
                        district_counts[d] = district_counts.get(d, 0) + 1
            else:
                # 普通省：只有市级归入本市市本级；省级应在省本级，不要混进来
                if level == "municipal":
                    muni_count += 1
                elif level == "county":
                    d = r[self.COL_DISTRICT]
                    if d:
                        district_counts[d] = district_counts.get(d, 0) + 1
        items.append(("市本级", muni_count, "MUNICIPAL"))
        for d, c in sorted(district_counts.items(), key=lambda x: -x[1]):
            items.append((d, c, d))
        return items

    def _national_status(self, bond_type, items):
        central = next((c for n, c, t in items if t == "CENTRAL"), 0)
        provinces = len([x for x in items if x[2] not in ("CENTRAL",)])
        return f"央企 {central} 家，{provinces} 个省份"

    def _province_status(self, province, items):
        if self._is_municipality(province):
            muni = next((c for n, c, t in items if t == "MUNICIPAL"), 0)
            districts = len([x for x in items if x[2] != "MUNICIPAL"])
            return f"市本级 {muni} 条，{districts} 个区县"
        else:
            prov = next((c for n, c, t in items if t == "PROVINCIAL"), 0)
            cities = len([x for x in items if x[2] != "PROVINCIAL"])
            return f"省本级 {prov} 条，{cities} 个地级市"

    def _city_status(self, items):
        muni = next((c for n, c, t in items if t == "MUNICIPAL"), 0)
        districts = len([x for x in items if x[2] != "MUNICIPAL"])
        return f"市本级 {muni} 条，{districts} 个区县"

    def _on_national_item_click(self, tag):
        bt = self.current_bond_type
        if tag == "CENTRAL":
            rows = self._central_rows(bt)
            self._show_detail("央企", rows)
        else:
            self._show_level(2, province=tag)

    def _on_province_item_click(self, tag):
        province = self.current_province
        bt = self.current_bond_type
        in_muni = self._is_municipality(province)
        if tag == "MUNICIPAL":
            if in_muni:
                rows = [r for r in self._filter_rows(bond_type=bt, province=province)
                        if self._row_level(r, province) in ("provincial", "municipal")]
            else:
                # 普通省份的“市本级”按钮不会出现，此处做兜底
                rows = [r for r in self._filter_rows(bond_type=bt, province=province)
                        if self._row_level(r, province) == "municipal"]
            self._show_detail(f"{province} 市本级", rows)
        elif tag == "PROVINCIAL":
            rows = [r for r in self._filter_rows(bond_type=bt, province=province)
                    if self._row_level(r, province) == "provincial"]
            self._show_detail(f"{province} 省本级", rows)
        else:
            if in_muni:
                # 直辖市：tag 是区/县名，直接显示该区县的区县级记录
                rows = [r for r in self._filter_rows(bond_type=bt, province=province, district=tag)
                        if self._row_level(r, province) == "county"]
                self._show_detail(f"{province} {tag}", rows)
            else:
                self._show_level(3, province=province, city=tag)

    def _on_city_item_click(self, tag):
        province = self.current_province
        city = self.current_city
        bt = self.current_bond_type
        in_muni = self._is_municipality(province)
        if tag == "MUNICIPAL":
            if in_muni:
                rows = [r for r in self._filter_rows(bond_type=bt, province=province, city=city)
                        if self._row_level(r, province) in ("provincial", "municipal")]
            else:
                # 普通地级市市本级只包含市级记录，不要混入省级
                rows = [r for r in self._filter_rows(bond_type=bt, province=province, city=city)
                        if self._row_level(r, province) == "municipal"]
            self._show_detail(f"{city} 市本级", rows)
        else:
            rows = [r for r in self._filter_rows(bond_type=bt, province=province, city=city, district=tag)
                    if self._row_level(r, province) == "county"]
            self._show_detail(f"{city} · {tag}", rows)

    # ------------------------------------------------------------------
    # 分类明细处理
    # ------------------------------------------------------------------
    def _switch_tab(self, tab_name):
        """切换明细分类 Tab，并高亮当前选中按钮。"""
        self.current_tab = tab_name
        for name, frame in self.tab_frames.items():
            if name == tab_name:
                frame.grid()
            else:
                frame.grid_remove()
        for name, btn in self.tab_buttons.items():
            if name == tab_name:
                btn.config(bg="#4a90d9", fg="white", relief="raised", bd=2)
            else:
                btn.config(bg="#f0f0f0", fg="#333333", relief="flat", bd=1)

    def _show_detail(self, subtitle, rows):
        self.nav_frame.grid_remove()
        self.detail_frame.grid()
        self.detail_title.config(
            text=f"{self.current_bond_type or ''} {subtitle} — 共 {len(rows)} 条"
        )
        self.title_lbl.config(text="发债主体信息")
        self.status_lbl.config(text="正在加载明细…")

        for frozen_tree, scroll_tree, _ in self.tab_trees.values():
            frozen_tree.delete(*frozen_tree.get_children())
            scroll_tree.delete(*scroll_tree.get_children())

        self._auto_fit_columns(rows)
        self._current_detail_rows = rows
        self._switch_tab("汇总")
        self._populate_tab_batch("汇总", rows, 0)

    def _populate_tab_batch(self, tab_name, rows, start):
        frozen_tree, scroll_tree, freeze_idx = self.tab_trees[tab_name]
        cols = self.TAB_COLUMNS[tab_name]
        end = min(start + self.INSERT_BATCH, len(rows))
        for r in rows[start:end]:
            values = []
            for _, src in cols:
                if callable(src):
                    try:
                        values.append(src(r))
                    except Exception:
                        values.append("")
                else:
                    values.append(r[src] if src < len(r) else "")
            frozen_tree.insert("", tk.END, values=[values[freeze_idx]])
            scroll_values = [v for i, v in enumerate(values) if i != freeze_idx]
            scroll_tree.insert("", tk.END, values=scroll_values)

        if end < len(rows):
            self.status_lbl.config(text=f"正在加载 {tab_name}… {end}/{len(rows)}")
            self.after(10, lambda: self._populate_tab_batch(tab_name, rows, end))
        else:
            tab_order = ["汇总", "主体信息", "债券信息", "财务信息", "第一大收入构成"]
            idx = tab_order.index(tab_name)
            if idx + 1 < len(tab_order):
                next_tab = tab_order[idx + 1]
                self.status_lbl.config(text=f"正在加载 {next_tab}…")
                self.after(10, lambda: self._populate_tab_batch(next_tab, self._current_detail_rows, 0))
            else:
                self.status_lbl.config(text=f"当前显示 {len(rows)} 条")

    def _go_back(self):
        if self.detail_frame.winfo_viewable():
            self.detail_frame.grid_remove()
            self.nav_frame.grid()
            if self.current_level == 0:
                self.nav_sub_frame.grid_remove()
                self.bond_type_frame.place(x=0, y=0)
                self.bond_type_frame.lift()
                self._center_bond_type_frame()
            else:
                self.bond_type_frame.place_forget()
                self.nav_sub_frame.grid()
            self._update_title()
            self.status_lbl.config(text="")
            return

        if self.current_level == 3:
            self._show_level(2, province=self.current_province)
        elif self.current_level == 2:
            self._show_level(1)
        elif self.current_level == 1:
            self._show_bond_type_selection()

    def _go_home(self):
        if self.current_bond_type:
            self._show_level(1)
        else:
            self._show_bond_type_selection()

    def _update_title(self):
        bt = self.current_bond_type
        if self.current_level == 1:
            self.title_lbl.config(text=f"{bt} — 全国省份/央企")
        elif self.current_level == 2:
            self.title_lbl.config(text=f"{bt} — {self.current_province} 下辖层级")
        elif self.current_level == 3:
            self.title_lbl.config(text=f"{bt} — {self.current_province} · {self.current_city} 下辖层级")

    def _on_close(self):
        self.destroy()
        sys.exit(0)

    # ------------------------------------------------------------------
    # 格式化
    # ------------------------------------------------------------------
    @staticmethod
    def _fmt_header(value):
        if value is None:
            return ""
        return str(value).replace("\n", " ").strip()

    @staticmethod
    def _fmt_value(value):
        if value is None:
            return ""
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value).strip()


def main():
    app = BondBrowserApp()
    app.mainloop()


if __name__ == "__main__":
    main()
